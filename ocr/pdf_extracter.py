"""
Demo script using Tesseract OCR.
Extract text of a page and interpret unrecognized characters using Tesseract.
MuPDF codes unrecognizable characters as 0xFFFD = 65533.
Extraction option is "dict", which delivers contiguous text pieces within one
line, that have the same font properties (color, fontsize, etc.). Together with
the language parameter, this helps Tesseract finding the correct character.
The basic approach is to only invoke OCR, if the span text contains
chr(65533). Because Tesseract's response ignores leading spaces and appends
line break characters, some adjustments are made.
--------------
This demo will OCR only text, that is known to be text. This means, it
does not look at parts of a page containing images or text encoded as drawings.
--------------
Dependencies:
Tesseract must be installed and invocable via Python's 'subprocess' module.
You also must install all the Tesseract language support you need to detect.
"""
import fitz
import subprocess
import time
from datetime import datetime
import glob
import re
import numpy as np
from numpy.core.arrayprint import format_float_scientific
import pandas as pd
from os import walk
import os
import glob
from pathlib import Path
from openpyxl import load_workbook
from shutil import copyfile
import requests
import pymysql
import pymysql.cursors
from json import loads
import csv
import json
from sqlalchemy import create_engine
from urllib.parse import quote


ENCODING = "UTF-8"
# PARRENT_PATH = "/home/data/Documents/Crawl_info"
# MY_PATH = "/home/data/Documents/Crawl_info/temp"
# CSV_PATH = "/home/data/Documents/Crawl_info/temp_csv"

PARRENT_PATH = "/Users/dinhvan/Documents/Projects/OCR/Crawl_info"
MY_PATH = "/Users/dinhvan/Documents/Projects/OCR/Crawl_info/temp"
CSV_PATH = "/Users/dinhvan/Documents/Projects/OCR/Crawl_info/temp_csv"

# MY_PATH = "/home/ptdl/Documents/Projects/Crawl_info/temp"
# CSV_PATH = "/home/ptdl/Documents/Projects/Crawl_info/temp_csv"

current_path = os.path.dirname(os.path.abspath(__file__))
khdn_quan_path = os.path.join(current_path, "khdn_data", "quan")
khdn_phuong_path = os.path.join(current_path, "khdn_data", "phuong")

path_to_file = glob.glob(MY_PATH + "/*/")
path_to_file.sort(key = lambda f: int(re.sub("\D", "", f)))
# path_to_file = []
# for (dirpath, dirnames, filenames) in walk(MY_PATH):
#     for dirname in dirnames:
#         path_to_file.append(dirpath + "/" + dirname)

sqlEngine = create_engine(
    "mysql+pymysql://root:%s@172.16.10.112:3306/hkd" % quote("Ptdl@123")
)

TYPE_KHDN = "khdn"

code_nganh_chinh = {
    "A": [str(i).zfill(2) for i in range(1, 4)],
    "B": [str(i).zfill(2) for i in range(5, 10)],
    "C": [str(i).zfill(2) for i in range(10, 34)],
    "D": [str(35).zfill(2)],
    "E": [str(i).zfill(2) for i in range(36, 40)],
    "F": [str(i).zfill(2) for i in range(41, 44)],
    "G": [str(i).zfill(2) for i in range(45, 48)],
    "H": [str(i).zfill(2) for i in range(49, 54)],
    "I": [str(i).zfill(2) for i in range(55, 57)],
    "J": [str(i).zfill(2) for i in range(58, 64)],
    "K": [str(i).zfill(2) for i in range(64, 67)],
    "L": [str(68).zfill(2)],
    "M": [str(i).zfill(2) for i in range(69, 76)],
    "N": [str(i).zfill(2) for i in range(77, 83)],
    "O": [str(84).zfill(2)],
    "P": [str(85).zfill(2)],
    "Q": [str(i).zfill(2) for i in range(86, 89)],
    "R": [str(i).zfill(2) for i in range(90, 94)],
    "S": [str(i).zfill(2) for i in range(94, 97)],
    "T": [str(i).zfill(2) for i in range(97, 99)],
    "U": [str(99).zfill(2)],
}


def matchingKeys(dictionary, searchString):
    return [
        key for key, val in dictionary.items() if any(searchString in s for s in val)
    ]
# Tesseract invocation command (Windows version)
# Assume: language English. Detect more languages by add e.g. '+deu' for German.
# Assume: text represents one line (--psm 7)
# Note: Language mix spec increases duration by >40% - only use when needed!
tess = "tesseract stdin stdout --oem 3 --psm 6 -l vie"
mat = fitz.Matrix(4, 4)  # high resolution matrix
ocr_time = 0
pix_time = 0

Ma_TTP = {
    "HNI": "TTKD Hà Nội",
    "VPC": "TTKD Vĩnh Phúc",
    "HBH": "TTKD Hòa Bình",
    "BNH": "TTKD Bắc Ninh",
    "BCN": "TTKD Bắc Kạn",
    "LCI": "TTKD Lào Cai",
    "LSN": "TTKD Lạng Sơn",
    "BGG": "TTKD Bắc Giang",
    "CBG": "TTKD Cao Bằng",
    "TNN": "TTKD Thái Nguyên",
    "PTO": "TTKD Phú Thọ",
    "TQG": "TTKD Tuyên Quang",
    "YBI": "TTKD Yên Bái",
    "SLA": "TTKD Sơn La",
    "DBN": "TTKD Điện Biên",
    "LCU": "TTKD Lai Châu",
    "HGG": "TTKD Hà Giang",
    "HNM": "TTKD Hà Nam",
    "NDH": "TTKD Nam Định",
    "TBH": "TTKD Thái Bình",
    "HDG": "TTKD Hải Dương",
    "HPG": "TTKD Hải Phòng",
    "QNH": "TTKD Quảng Ninh",
    "HYN": "TTKD Hưng Yên",
    "NBH": "TTKD Ninh Bình",
    "THA": "TTKD Thanh Hoá",
    "NAN": "TTKD Nghệ An",
    "HTH": "TTKD Hà Tĩnh",
    "QBH": "TTKD Quảng Bình",
    "QTI": "TTKD Quảng Trị",
    "HUE": "TTKD Thừa Thiên-Huế",
    "QNM": "TTKD Quảng Nam",
    "QNI": "TTKD Quảng Ngãi",
    "BDH": "TTKD Bình Định",
    "GLI": "TTKD Gia Lai",
    "DLC": "TTKD Đắk Lắk",
    "DKN": "TTKD Đắk Nông",
    "PYN": "TTKD Phú Yên",
    "KHA": "TTKD Khánh Hòa",
    "KTM": "TTKD KonTum",
    "DNG": "TTKD Đà Nẵng",
    "LDG": "TTKD Lâm Đồng",
    "BTN": "TTKD Bình Thuận",
    "NTN": "TTKD Ninh Thuận",
    "HCM": "TTKD TP Hồ Chí Minh",
    "DNI": "TTKD Đồng Nai",
    "BDG": "TTKD Bình Dương",
    "TNH": "TTKD Tây Ninh",
    "VTU": "TTKD Bà Rịa - Vũng Tàu",
    "BPC": "TTKD Bình Phước",
    "LAN": "TTKD Long An",
    "TGG": "TTKD Tiền Giang",
    "BTE": "TTKD Bến Tre",
    "TVH": "TTKD Trà Vinh",
    "VLG": "TTKD Vĩnh Long",
    "CTO": "TTKD Cần Thơ",
    "HAG": "TTKD Hậu Giang",
    "DTP": "TTKD Đồng Tháp",
    "AGG": "TTKD An Giang",
    "KGG": "TTKD Kiên Giang",
    "CMU": "TTKD Cà Mau",
    "STG": "TTKD Sóc Trăng",
    "BLU": "TTKD Bạc Liêu",
}

Vung_TTP = {
    "MB": "Lào Cai,  Điện Biên, Hòa Bình, Lai Châu, Sơn La, Hà Giang, Cao Bằng, Bắc Kạn, Lạng Sơn, Tuyên Quang, Thái Nguyên, Phú Thọ, Bắc Giang, Quảng Ninh, Bắc Ninh, Hà Nam, Hà Nội, Hải Dương, Thanh Hoá, Hưng Yên,  Nam Định, Thái Bình, Vĩnh Phúc",
    "MT": "Yên Bái, Nghệ An, Ninh Bình, Tuyên Quang, Hà Tĩnh , Quảng Bình,  Quảng Trị, Thừa Thiên-Huế, Đà Nẵng, Quảng Nam, Quảng Ngãi, Bình Định, Phú Yên, Khánh Hòa, KonTum, Gia Lai, Đắk Lắk, Đắk Nông, Hải Phòng",
    "MN": "Bình Phước, Ninh Thuận, Bình Thuận, Bình Dương, Đồng Nai, Tây Ninh, Bà Rịa-Vũng Tàu, Thành phố Hồ Chí Minh, Long An, Đồng Tháp, Tiền Giang, An Giang, Bến Tre, Vĩnh Long, Trà Vinh, Hậu Giang, Kiên Giang, Sóc Trăng, Bạc Liêu, Cà Mau, Thành phố Cần Thơ, Lâm Đồng",
}

Khdn_tinh_id = {
    'an giang':'1',
    'bình dương':'8',
    'bình phước':'9',
    'bình thuận':'10',
    'bình định':'6',
    'bạc liêu':'11',
    'bắc cạn':'4',
    'bắc giang':'3',
    'bắc ninh':'5',
    'bến tre':'7',
    'cao bằng':'12',
    'cà mau':'14',
    'cần thơ':'13',
    'gia lai':'19',
    'huế':'53',
    'hà giang':'20',
    'hà nam':'25',
    'hà nội':'21',
    'hà tĩnh':'23',
    'hòa bình':'65',
    'hưng yên':'24',
    'hải dương':'27',
    'hải phòng':'26',
    'hậu giang':'66',
    'khánh hoà':'29',
    'không xác định':'99',
    'kiên giang':'30',
    'kon tum':'31',
    'lai châu':'32',
    'long an':'36',
    'lào cai':'34',
    'lâm đồng':'35',
    'lạng sơn':'33',
    'nam định':'37',
    'net':'67',
    'nghệ an':'38',
    'ninh bình':'39',
    'ninh thuận':'40',
    'phú thọ':'59',
    'phú yên':'41',
    'quảng bình':'42',
    'quảng nam':'43',
    'quảng ngãi':'44',
    'quảng ninh':'45',
    'quảng trị':'46',
    'quốc tế (vti)':'98',
    'sóc trăng':'47',
    'sơn la':'49',
    'chí minh':'28',
    'đà nẵng':'15',
    'thanh hoá':'52',
    'thái bình':'51',
    'thái nguyên':'61',
    'tiền giang':'54',
    'trà vinh':'55',
    'tuyên quang':'56',
    'tây ninh':'50',
    'điện biên':'22',
    'vinaphone':'100',
    'vĩnh long':'57',
    'vĩnh phúc':'58',
    'vũng tàu':'2',
    'yên bái':'60',
    'đắk lắk':'16',
    'đắk nông':'64',
    'đồng nai':'17',
    'đồng tháp':'18',
}
Khdn_tinh_id = {
    "HNI": "21",
    "VPC": "58",
    "HBH": "65",
    "BNH": "5",
    "BCN": "4",
    "LCI": "34",
    "LSN": "33",
    "BGG": "3",
    "CBG": "12",
    "TNN": "61",
    "PTO": "59",
    "TQG": "56",
    "YBI": "60",
    "SLA": "49",
    "DBN": "22",
    "LCU": "32",
    "HGG": "20",
    "HNM": "25",
    "NDH": "37",
    "TBH": "51",
    "HDG": "27",
    "HPG": "26",
    "QNH": "45",
    "HYN": "24",
    "NBH": "39",
    "THA": "52",
    "NAN": "38",
    "HTH": "23",
    "QBH": "42",
    "QTI": "46",
    "HUE": "53",
    "QNM": "43",
    "QNI": "44",
    "BDH": "6",
    "GLI": "19",
    "DLC": "16",
    "DKN": "64",
    "PYN": "41",
    "KHA": "29",
    "KTM": "31",
    "DNG": "15",
    "LDG": "35",
    "BTN": "10",
    "NTN": "40",
    "HCM": "28",
    "DNI": "17",
    "BDG": "8",
    "TNH": "50",
    "VTU": "2",
    "BPC": "9",
    "LAN": "36",
    "TGG": "54",
    "BTE": "7",
    "TVH": "55",
    "VLG": "57",
    "CTO": "13",
    "HAG": "66",
    "DTP": "18",
    "AGG": "1",
    "KGG": "30",
    "CMU": "14",
    "STG": "47",
    "BLU": "11",
}

Phone_Prefix_Vina = {
    "096": "Viettel",
    "097": "Viettel",
    "098": "Viettel",
    "032": "Viettel",
    "033": "Viettel",
    "034": "Viettel",
    "035": "Viettel",
    "036": "Viettel",
    "037": "Viettel",
    "038": "Viettel",
    "039": "Viettel",
    "086": "Viettel",
    "090": "Mobifone",
    "093": "Mobifone",
    "070": "Mobifone",
    "089": "Mobifone",
    "077": "Mobifone",
    "076": "Mobifone",
    "078": "Mobifone",
    "079": "Mobifone",
    "091": "Vinaphone",
    "094": "Vinaphone",
    "081": "Vinaphone",
    "082": "Vinaphone",
    "083": "Vinaphone",
    "084": "Vinaphone",
    "085": "Vinaphone",
    "088": "Vinaphone",
    "099": "Gmobile",
    "059": "Gmobile",
    "092": "Vietnamobile",
    "056": "Vietnamobile",
    "058": "Vietnamobile",
}

def conn():
    MYSQL_HOST = "172.16.10.112"
    MYSQL_PORT = 3306
    MYSQL_DBNAME = "hkd"
    MYSQL_USER = "root"
    MYSQL_PASSWD = "Ptdl@123"

    connection = pymysql.connect(
        host=MYSQL_HOST,
        user=MYSQL_USER,
        password=MYSQL_PASSWD,
        database=MYSQL_DBNAME,
        port=MYSQL_PORT,
        charset="utf8mb4",
        autocommit=True,
        cursorclass=pymysql.cursors.DictCursor,
    )
    return connection

def get_nganh_nghe_by_id(id):
    result = "Khác"
    if 1 <= id <= 3:
        result = "Nông nghiệp"
    elif 49 <= id <= 53:
        result = "Vận tải và Logictic"
    elif 5 <= id <= 43:
        result = "Công nghiệp và xây dựng"
    elif 86 <= id <= 88:
        result = "Y tế"
    elif id == 85:
        result = "Giáo dục"
    elif 45 <= id <= 47:
        result = "Phân phối, bán lẻ"
    elif id == 55 or id == 56 or id == 79:
        result = "Du lịch"
    return result

def append_df_to_excel(
    filename,
    df,
    sheet_name = "Sheet1",
    startrow = None,
    truncate_sheet = False,
    **to_excel_kwargs
):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name = sheet_name,
            startrow = startrow if startrow is not None else 0,
            **to_excel_kwargs
        )
        return

    # ignore [engine] parameter if it was passed
    if "engine" in to_excel_kwargs:
        to_excel_kwargs.pop("engine")

    writer = pd.ExcelWriter(filename, engine="openpyxl", mode="a")

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

def sort_blocks(blocks):
    """
    Sort the blocks of a TextPage in ascending vertical pixel order,
    then in ascending horizontal pixel order.
    This should sequence the text in a more readable form, at least by
    convention of the Western hemisphere: from top-left to bottom-right.
    If you need something else, change the sortkey variable accordingly ...
    """

    sblocks = []
    for b in blocks:
        x0 = str(int(b["bbox"][0] + 0.99999)).rjust(4, "0")  # x coord in pixels
        y0 = str(int(b["bbox"][1] + 0.99999)).rjust(4, "0")  # y coord in pixels
        sortkey = y0 + x0  # = "yx"
        sblocks.append([sortkey, b])
    sblocks.sort()
    return [b[1] for b in sblocks]  # return sorted list of blocks

def sort_lines(lines):
    """Sort the lines of a block in ascending vertical direction. See comment
    in sort_blocks function.
    """
    slines = []
    for l in lines:
        y0 = str(int(l["bbox"][1] + 0.99999)).rjust(4, "0")
        slines.append([y0, l])
    slines.sort()
    return [l[1] for l in slines]

def sort_spans(spans):
    """Sort the spans of a line in ascending horizontal direction. See comment
    in sort_blocks function.
    """
    sspans = []
    for s in spans:
        x0 = str(int(s["bbox"][0] + 0.99999)).rjust(4, "0")
        sspans.append([x0, s])
    sspans.sort()
    return [s[1] for s in sspans]

def get_tessocr(page, bbox):
    """Return OCR-ed span text using Tesseract.
    Args:
        page: fitz.Page
        bbox: fitz.Rect or its tuple
    Returns:
        The OCR-ed text of the bbox.
    """
    global ocr_time, pix_time, tess, mat
    # Step 1: Make a high-resolution image of the bbox.
    
    t0 = time.perf_counter()
    pix = page.get_pixmap(
        colorspace  =fitz.csGRAY,  # we need no color
        matrix      = mat,
        clip        = bbox,
    )
    image = pix.getImageData("png")  # make a PNG image
    
    t1 = time.perf_counter()
    # Step 2: Invoke Tesseract to OCR the image. Text is stored in stdout.
    rc = subprocess.run(
        tess,  # the command
        input   =image,  # the pixmap image
        stdout  =subprocess.PIPE,  # find the text here
        shell   =True,
    )
    # because we told Tesseract to interpret the image as one line, we now need
    # to strip off the line break characters from the tail.
    text = rc.stdout.decode()  # convert to string
    text = re.sub(r"\n\x0c", "", text)  # remove line end characters
    
    t2 = time.perf_counter()
    ocr_time += t2 - t1
    pix_time += t1 - t0
    return text

def parse_pdf(path):
    results = []
    try:
        doc = fitz.open(path)
        ocr_count = 0
        for page in doc:
            page_get_text = page.get_text("dict", flags=0)
            blocks = sort_blocks(page_get_text["blocks"])
            for b in blocks:
                # lines = SortLines(b["lines"])            # ... lines
                for l in b["lines"]:
                    spans = sort_spans(l["spans"])  # ... spans
                    for s in spans:
                        text = s["text"]
                        if chr(65533) in text:  # invalid characters encountered!
                            # invoke OCR
                            ocr_count += 1
                            # print("before: '%s'" % text)
                            text1 = text.lstrip()
                            sb = " " * (len(text) - len(text1))  # leading spaces
                            text1 = text.rstrip()
                            sa = " " * (len(text) - len(text1))  # trailing spaces
                            text = sb + get_tessocr(page, s["bbox"]) + sa
                            # print(" after: '%s'" % text)
                        # else:
                        # print("non ocr: '%s'" % text)
                        # if "6." in text and ":" in text:
                        #     return results
                        # else:
                        results.append(text)

        # print("-------------------------")
        # print("OCR invocations: %i." % ocr_count)
        # print(
        #     "Pixmap time: %g (avg %g) seconds."
        #     % (round(pix_time, 5), round(pix_time / ocr_count, 5))
        # )
        # print(
        #     "OCR time: %g (avg %g) seconds."
        #     % (round(ocr_time, 5), round(ocr_time / ocr_count, 5))
        # )
    except Exception as e:
        print(e)
        
    return results

def pdf_to_array(index_titles, info_arrays):
    info_company = []
    last_info_index = len(info_arrays)
    index_titles.append(last_info_index)
    for key, index in enumerate(index_titles):
        temp_info = ""
        if key < len(index_titles) - 1:
            slide = range(index, index_titles[key + 1])
            for i in slide:
                if "5. " in info_arrays[i]:
                    nn_temp_company = ""
                    nganh_nghe_index_array = index_titles[key + 1 :]
                    for nn_key, nn_value in enumerate(nganh_nghe_index_array):
                        nn_temp_info = ""
                        if nn_key < len(nganh_nghe_index_array) - 1:
                            for j in range(
                                nganh_nghe_index_array[nn_key],
                                nganh_nghe_index_array[nn_key + 1],
                            ):
                                nn_temp_info += " " + info_arrays[j]
                        # elif nn_key == len(nganh_nghe_index_array) - 1:
                        #     last_index = nganh_nghe_index_array[nn_key]
                        #     for info in info_arrays[last_index:]:
                        #         nn_temp_info += "_" + info
                        nn_temp_company += nn_temp_info.lstrip("_") + "\n"
                    info_company.append(nn_temp_company)
                    return info_company
                else:
                    if ":" in info_arrays[i]:
                        temp_info += "\n" + info_arrays[i]
                    else:
                        temp_info += " " + info_arrays[i]
            info_company.append(temp_info.lstrip("\n"))
    return info_company

def preprocess_info(str, is_nn):
    result = ""
    if "\n" in str:
        str = str.split("\n")[1:]
        for s in str:
            result += s + "\n"
    else:
        str = str.split(".")[1:]
        for s in str:
            result += s

    result = result.lstrip(": _")
    # result = units.convert_unicode(result)
    # result = units.chuan_hoa_dau_cau_tieng_viet(result)
    # if not is_nn:
    #     temp_result = ""
    #     for x in result.split("_")[1:]:
    #         temp_result += x + "_"
    #     result = temp_result
    #     result = result.strip(': _')

    return result

def recover_info_company(info_company_arrays):
    ica = []
    ica_nn = []
    for key, info in enumerate(info_company_arrays):
        # if type(info) == list:
        #     for i in info:
        #         i = preprocess_info(i, True)
        #         ica_nn.append(i)
        #     ica.append(ica_nn)
        #     return ica
        if key != len(info_company_arrays) - 1:
            info = preprocess_info(info, False)
        ica.append(info)
    return ica

def parse_ocr_data(info_company_recovered,info_director_arrays,ocr_dir,pre_csv_data,report_fdate_path,report_tourist_path):
    gioi_tinh_key_array = [i for i, j in enumerate(info_director_arrays) if "Gới tính" in j or "tính:" in j][0]
    if info_company_recovered[3] == "\n":
        info_company_recovered.pop(3)
    if len(info_company_recovered) != 6:
        ma_so_thue = "'" + info_company_recovered[1].rstrip("\n")
        detail_info_director = info_company_recovered[3].split("\n")
        nganhnghe_info = info_company_recovered[4].split("\n")
        ngay_thanh_lap_info = info_company_recovered[2].split(":")[-1].lstrip()
    else:
        if "Ngày thành lập" in info_company_recovered[2]:
            ma_so_thue = "'" + info_company_recovered[1].rstrip("\n")
            ngay_thanh_lap_info = info_company_recovered[2].split(":")[-1].lstrip()
        else:
            created_date_mst = info_company_recovered[2].split(" ")
            for cdm in created_date_mst:
                if "/" in cdm:
                    ngay_thanh_lap_info = cdm
                else:
                    ma_so_thue = cdm.rstrip("\n")
        detail_info_director = info_company_recovered[3].split("\n")
        nganhnghe_info = info_company_recovered[-1].split("\n")

    nn_id = []
    for nn_info in nganhnghe_info:
        nn_id.append(nn_info.split(" ")[-1][0:2])
    try:

        # dia_chi_kh = detail_info_director[0]
        # ngay_cap = info_director_arrays[gioi_tinh_key_array+17]
        # noi_cap = info_director_arrays[gioi_tinh_key_array+18]
        # ho_khau = ''
        # so_dt = detail_info_director[3].split(":")[1]
        # email = detail_info_director[4].split(":")[1]
        # nguoi_dd = info_director_arrays[gioi_tinh_key_array+1].split(":")[-1].lstrip(" ")
        # noicap_hk = ''
        # ngay_hk = ''
        # ngay_sn = info_director_arrays[gioi_tinh_key_array+8]
        # dienthoai_lh = ''
        # gioi_tinh = info_director_arrays[gioi_tinh_key_array+2]
        nganhnghe_id = info_company_recovered[4]

        stt_row = pre_csv_data["stt_index"]
        province_code = pre_csv_data["province_code"]
        province_zip = pre_csv_data["province_zip"]

        df_check = pd.read_csv(ocr_dir)
        mst_check = df_check["Mã số thuế"].tolist()
        if ma_so_thue not in mst_check:

            ttdl_url = (
                "https://thongtindoanhnghiep.co/api/company/"
                + info_company_recovered[1].rstrip("\n")
            )
            payload = {}
            headers = {}

            dia_chi_tinh_tp = ""
            dia_chi_quan_huyen = ""
            dia_chi_phuong_xa = ""
            dia_chi_cu_the = ""

            try:
                ttdn_res = requests.request(
                    "GET", ttdl_url, headers=headers, data=payload
                )
                ttdn_info = ttdn_res.json()
                ten_tieng_viet = ttdn_info["Title"]
                ten_tieng_anh = ttdn_info["TitleEn"]
                dia_chi = ttdn_info["DiaChiCongTy"].split(",")
                if len(dia_chi) > 1:
                    dia_chi_tinh_tp = dia_chi[-1].lstrip()
                if len(dia_chi) > 2:
                    dia_chi_quan_huyen = dia_chi[-2].lstrip()
                if len(dia_chi) > 3:
                    dia_chi_phuong_xa = dia_chi[-3].lstrip()
                    dia_chi_cu_the = " ".join(dia_chi[:-3])
            except Exception as e:
                print(e)
                ten_khach_hang = info_company_recovered[0].split("\n")
                ten_tieng_viet = ten_khach_hang[0].split(":")[-1].lstrip()
                ten_tieng_anh = ten_khach_hang[1].split(":")[-1].lstrip()
                dia_chi = detail_info_director[0].split(",")
                if len(dia_chi) > 2:
                    dia_chi_tinh_tp = dia_chi[-2].lstrip()
                if len(dia_chi) > 3:
                    dia_chi_quan_huyen = dia_chi[-3].lstrip()
                if len(dia_chi) > 4:
                    dia_chi_phuong_xa = dia_chi[-4].lstrip()
                dia_chi_cu_the = detail_info_director[0]

            dien_thoai = (
                detail_info_director[3]
                .split(":")[1]
                .lstrip(" 0")
                .replace(" ", "")
                .replace(".", "")
            )
            if not dien_thoai.isnumeric() or dien_thoai == "":
                for detail in detail_info_director:
                    if "Điện thoại" in detail:
                        dien_thoai = (
                            detail.split(":")[1]
                            .lstrip(" 0")
                            .replace(" ", "")
                            .replace(".", "")
                        )
            if "/" in dien_thoai or "-" in dien_thoai or "–" in dien_thoai:
                if "/" in dien_thoai:
                    dien_thoai = dien_thoai.split("/")
                    dien_thoai_list = list()
                    for dt in dien_thoai:
                        dien_thoai_list.append("84{}".format(dt))
                    dien_thoai = "\n".join(dien_thoai_list)
                elif "–" in dien_thoai:
                    dien_thoai = dien_thoai.split("–")
                    if len(dien_thoai) > 2:
                        if isinstance(dien_thoai, list):
                            dien_thoai = "".join(dien_thoai)
                        else:
                            dien_thoai.replace("–", "")
                    else:
                        dien_thoai_list = list()
                        for dt in dien_thoai:
                            dien_thoai_list.append("84{}".format(dt))
                        dien_thoai = "\n".join(dien_thoai_list)
                elif "-" in dien_thoai:
                    dien_thoai = dien_thoai.split("-")
                    if len(dien_thoai) > 2:
                        if isinstance(dien_thoai, list):
                            dien_thoai = "".join(dien_thoai)
                        else:
                            dien_thoai.replace("-", "")
                    else:
                        dien_thoai_list = list()
                        for dt in dien_thoai:
                            dien_thoai_list.append("84{}".format(dt))
                        dien_thoai = "\n".join(dien_thoai_list)
            else:
                dien_thoai = "84{}".format(dien_thoai)

            if "." in dien_thoai:
                dien_thoai = dien_thoai.replace(".", "")

            prefix_phone = dien_thoai[:4]
            telco_name = "khác"
            try:
                telco_name = Phone_Prefix_Vina.get(prefix_phone)
            except KeyError as e:
                print(e)

            email = ""
            for detail in detail_info_director:
                if "@" in detail:
                    email = re.findall(
                        r"[\w\.-]+@[\w\.-]+", detail_info_director[4].split(":")[1]
                    )[0]
                    break

            nganh_nghe = list(
                set(
                    [
                        get_nganh_nghe_by_id(int(id))
                        for id in nn_id
                        if id.isnumeric() and id != ""
                    ]
                )
            )
            linh_vuc = "\n".join(nganh_nghe)

            name_chu_so_huu = ""
            dob_chu_so_huu = ""
            giay_to_chu_so_huu = ""
            dia_chi_chu_so_huu = ""
            lien_lac_chu_so_huu = ""
            dt_chu_so_huu = ""
            email_chu_so_huu = ""

            name_dai_dien = (
                info_director_arrays[gioi_tinh_key_array + 1].split(":")[-1].lstrip(" ")
            )

            if (gioi_tinh_key_array + 8) < len(info_director_arrays):
                dob_dai_dien = info_director_arrays[gioi_tinh_key_array + 8]
            else:
                dob_dai_dien = ""

            if (gioi_tinh_key_array + 14) < len(info_director_arrays):
                giay_to_dai_dien = info_director_arrays[gioi_tinh_key_array + 14]
            else:
                giay_to_dai_dien = ""

            if (gioi_tinh_key_array + 25) < len(info_director_arrays):
                dia_chi_dai_dien = info_director_arrays[gioi_tinh_key_array + 25]
            else:
                dia_chi_dai_dien = ""

            dia_chi_lien_lac = ""
            dt_dai_dien = ""
            email_dai_dien = ""

            obtt_name = ten_tieng_viet
            if not ten_tieng_viet:
                obtt_name = ten_tieng_anh

            obtt_dien_thoai = ""
            if "\n" in dien_thoai:
                obtt_dien_thoai = dien_thoai.split("\n")
                if len(obtt_dien_thoai[0]) > 2:
                    obtt_dien_thoai = obtt_dien_thoai[0][2:]
            elif "–" in dien_thoai:
                obtt_dien_thoai = dien_thoai.split("–")
                if len(obtt_dien_thoai[0]) > 2:
                    obtt_dien_thoai = obtt_dien_thoai[0][2:]
            elif "-" in dien_thoai:
                obtt_dien_thoai = dien_thoai.split("-")
                if len(obtt_dien_thoai[0]) > 2:
                    obtt_dien_thoai = obtt_dien_thoai[0][2:]
            elif ";" in dien_thoai:
                obtt_dien_thoai = dien_thoai.split(";")
                if len(obtt_dien_thoai[0]) > 2:
                    obtt_dien_thoai = obtt_dien_thoai[0][2:]
            else:
                obtt_dien_thoai = dien_thoai[2:]

            ocr_csv_data = np.array(
                [
                    [
                        dien_thoai,
                        telco_name,
                        obtt_name,
                        dia_chi_cu_the,
                        province_code,
                        dia_chi_quan_huyen,
                        dia_chi_phuong_xa,
                        ma_so_thue,
                        linh_vuc,
                        email,
                        ngay_thanh_lap_info,
                        name_dai_dien,
                        dob_dai_dien,
                        giay_to_dai_dien,
                        dia_chi_dai_dien,
                        province_zip,
                    ]
                ]
            )

            # obtt_name = ten_tieng_viet
            # if not ten_tieng_viet:
            #     obtt_name = ten_tieng_anh
            ocr_obtt_data = np.array(
                [
                    [
                        stt_row,
                        obtt_dien_thoai,
                        telco_name,
                        obtt_name,
                        dia_chi_cu_the,
                        province_code,
                        dia_chi_quan_huyen,
                        dia_chi_phuong_xa,
                        ma_so_thue,
                        linh_vuc,
                        email,
                        ngay_thanh_lap_info,
                        name_dai_dien,
                        dob_dai_dien,
                        giay_to_dai_dien,
                        dia_chi_dai_dien,
                        province_zip,
                    ]
                ]
            )

            """ EXPORT DATA THEO TUNG TINH """

            if not os.path.exists(ocr_dir):
                # csv_header = np.array([["KHACHHANG_ID", "MA_KH", "TEN_KH", "DIACHI_KH", "LOAIKH_ID", "KHLON_ID", "NGAYLAP_HD", "SO_GT", "NGAYCAP", "NOICAP", "HOKHAU", "SO_DT", "EMAIL", "NGUOI_DD", "NOICAP_HK", "NGAY_HK", "NGAY_SN", "DIENTHOAI_LH", "GIOITINH", "NGANHNGHE_ID", "TNC1_ID", "TNC2_ID", "TNC3_ID", "KHACHHANGCU_ID", "MA_HD", "PHANLOAIKH_ID", "LOAI", "DIEMTINNHIEM", "LOAIGT_ID", "MST"]])

                csv_header = np.array(
                    [
                        [
                            "Số thuê bao",
                            "Loại TB",
                            "Tên doanh nghiệp",
                            "Địa chỉ",
                            "Mã Tỉnh/TP",
                            "Quận/Huyện",
                            "Phường/Xã",
                            "Mã số thuế",
                            "Ngành nghề DN",
                            "Email",
                            "Ngày thành lập",
                            "Tên người đại diện",
                            "Ngày sinh",
                            "Số giấy tờ",
                            "Địa chỉ thường trú",
                            "Miền",
                        ]
                    ]
                )
                df = pd.DataFrame(data=csv_header)
                df.to_csv(ocr_dir, index=False, header=False)

                df_data = pd.DataFrame(data=ocr_csv_data)
                df_data.to_csv(
                    ocr_dir, index=False, header=False
                )
            else:
                df = pd.DataFrame(data=ocr_csv_data)
                df.to_csv(
                    ocr_dir,
                    index=False,
                    header=False,
                    mode="a",
                )

            """ EXPORT DATA BAO CAO NGAY """

            if not os.path.exists(report_fdate_path):
                # csv_header = np.array([["Vùng", "Mã T/TP", "Tên Tiếng Việt", "Tên Tiếng Anh", "Mã số thuế/Mã ngân sách", "Ngày thành lập DN", "Tỉnh/TP", "Quận/Huyện", "Phường/Xã", "Địa chỉ cụ thể ", "Điện thoại", "CHECK VINA", "Email", "Lĩnh vực (Ngành nghề kinh doanh chính)", "Họ và tên", "Ngày sinh", "Số giấy tờ", "Địa chỉ thường trú", "Địa chỉ liên lạc", "Số điện thoại", "Email", "Họ và tên", "Ngày sinh", "Số giấy tờ", "Địa chỉ thường trú", "Địa chỉ liên lạc", "Số điện thoại", "Email"]])
                # csv_ddate = np.concatenate((csv_header, ocr_csv_data))

                csv_header = np.array(
                    [
                        [
                            "STT",
                            "Số thuê bao",
                            "Loại TB",
                            "Tên doanh nghiệp",
                            "Địa chỉ",
                            "Mã Tỉnh/TP",
                            "Quận/Huyện",
                            "Phường/Xã",
                            "Mã số thuế",
                            "Ngành nghề DN",
                            "Email",
                            "Ngày thành lập",
                            "Tên người đại diện",
                            "Ngày sinh",
                            "Số giấy tờ",
                            "Địa chỉ thường trú",
                            "Miền",
                        ]
                    ]
                )
                csv_ddate = np.concatenate((csv_header, ocr_obtt_data))
                df = pd.DataFrame(data=csv_ddate)
                df.to_csv(
                    report_fdate_path,
                    index=False,
                    header=False,
                    sep="\t",
                    encoding="utf-8",
                )
            else:
                df = pd.DataFrame(data=ocr_obtt_data)
                df.to_csv(
                    report_fdate_path,
                    index=False,
                    header=False,
                    mode="a",
                    sep="\t",
                    encoding="utf-8",
                )

            """ EXPORT DATA TOURIST FOR MKT """

            # if "Du lịch" in nganh_nghe:
            #    if not os.path.exists(report_tourist_path):
            #        csv_header = np.array(
            #            [
            #                [
            #                    "STT",
            #                    "Số thuê bao",
            #                    "Loại TB",
            #                    "Tên doanh nghiệp",
            #                    "Địa chỉ",
            #                    "Mã Tỉnh/TP",
            #                    "Quận/Huyện",
            #                    "Phường/Xã",
            #                    "Mã số thuế",
            #                    "Ngành nghề DN",
            #                    "Email",
            #                    "Ngày thành lập",
            #                    "Tên người đại diện",
            #                    "Ngày sinh",
            #                    "Số giấy tờ",
            #                    "Địa chỉ thường trú",
            #                    "Miền",
            #                ]
            #            ]
            #        )
            #        csv_ddate = np.concatenate((csv_header, ocr_obtt_data))
            #        df = pd.DataFrame(data=csv_ddate)
            #        df.to_csv(report_tourist_path, index=False, header=False)
            #    else:
            #        df = pd.DataFrame(data=ocr_obtt_data)
            #        df.to_csv(report_tourist_path, index=False, header=False, mode="a")

            """ EXPORT DATA ALL PTDL """
            ocr_all_data = np.array(
                [
                    [
                        province_zip,
                        province_code,
                        ten_tieng_viet,
                        ten_tieng_anh,
                        ma_so_thue,
                        ngay_thanh_lap_info,
                        dia_chi_tinh_tp,
                        dia_chi_quan_huyen,
                        dia_chi_phuong_xa,
                        dia_chi_cu_the,
                        dien_thoai,
                        telco_name,
                        email,
                        nganhnghe_id,
                        name_chu_so_huu,
                        dob_chu_so_huu,
                        giay_to_chu_so_huu,
                        dia_chi_chu_so_huu,
                        lien_lac_chu_so_huu,
                        dt_chu_so_huu,
                        email_chu_so_huu,
                        name_dai_dien,
                        dob_dai_dien,
                        giay_to_dai_dien,
                        dia_chi_dai_dien,
                        dia_chi_lien_lac,
                        dt_dai_dien,
                        email_dai_dien,
                    ]
                ]
            )

            # if not os.path.exists(all_report_ptdl_path):
            #    csv_header = np.array(
            #        [
            #            [
            #                "Vùng",
            #                "Mã T/TP",
            #                "Tên Tiếng Việt",
            #                "Tên Tiếng Anh",
            #                "Mã số thuế/Mã ngân sách",
            #                "Ngày thành lập DN",
            #                "Tỉnh/TP",
            #                "Quận/Huyện",
            #                "Phường/Xã",
            #                "Địa chỉ cụ thể ",
            #                "Điện thoại",
            #                "CHECK VINA",
            #                "Email",
            #                "Lĩnh vực (Ngành nghề kinh doanh chính)",
            #                "Họ và tên",
            #                "Ngày sinh",
            #                "Số giấy tờ",
            #                "Địa chỉ thường trú",
            #                "Địa chỉ liên lạc",
            #                "Số điện thoại",
            #                "Email",
            #                "Họ và tên",
            #                "Ngày sinh",
            #                "Số giấy tờ",
            #                "Địa chỉ thường trú",
            #                "Địa chỉ liên lạc",
            #                "Số điện thoại",
            #                "Email",
            #            ]
            #        ]
            #    )
            #    csv_all_data = np.concatenate((csv_header, ocr_all_data))
            #    df = pd.DataFrame(data=csv_all_data)
            #    df.to_csv(all_report_ptdl_path, index=False, header=False)
            # else:
            #    df = pd.DataFrame(data=ocr_all_data)
            #    df.to_csv(all_report_ptdl_path, index=False, header=False, mode="a")

            # tạo file temp lưu dữ liệu đầy đủ theo ngày
            if not os.path.exists(os.path.join(PARRENT_PATH, "temp_ptdl.csv")):
                csv_header = np.array(
                    [
                        [
                            "Vùng",
                            "Mã T/TP",
                            "Tên Tiếng Việt",
                            "Tên Tiếng Anh",
                            "Mã số thuế/Mã ngân sách",
                            "Ngày thành lập DN",
                            "Tỉnh/TP",
                            "Quận/Huyện",
                            "Phường/Xã",
                            "Địa chỉ cụ thể ",
                            "Điện thoại",
                            "CHECK VINA",
                            "Email",
                            "Lĩnh vực (Ngành nghề kinh doanh chính)",
                            "Họ và tên",
                            "Ngày sinh",
                            "Số giấy tờ",
                            "Địa chỉ thường trú",
                            "Địa chỉ liên lạc",
                            "Số điện thoại",
                            "Email",
                            "Họ và tên",
                            "Ngày sinh",
                            "Số giấy tờ",
                            "Địa chỉ thường trú",
                            "Địa chỉ liên lạc",
                            "Số điện thoại",
                            "Email",
                        ]
                    ]
                )
                csv_all_data = np.concatenate((csv_header, ocr_all_data))
                df = pd.DataFrame(data=csv_all_data)
                df.to_csv(
                    os.path.join(PARRENT_PATH, "temp_ptdl.csv"),
                    index=False,
                    header=False,
                    encoding="utf-8",
                )
            else:
                df = pd.DataFrame(data=ocr_all_data)
                df.to_csv(
                    os.path.join(PARRENT_PATH, "temp_ptdl.csv"),
                    index=False,
                    header=False,
                    mode="a",
                    encoding="utf-8",
                )

            """ Add info dn to db """
            # ngay_thanh_lap_info = datetime.strptime(ngay_thanh_lap_info, "%d/%m/%Y")
            # timestamp = datetime.timestamp(ngay_thanh_lap_info)
            # datetime_temp = datetime.fromtimestamp(timestamp)
            # ngay_thanh_lap_info = datetime_temp.strftime("%Y-%m-%d %H:%M:%S")

            # Get nganh nghe chinh
        # nganh_nghe_list = nganhnghe_id.split("\n")
        # search_nn = [x for x in nganh_nghe_list if "(chính)" in x.lower()]
        # nganh_nghe_chinh = "\n".join(search_nn)
        # main_business_code = ""
        # if len(search_nn) > 0:
        #     nganh_chinh_first = search_nn[0]
        #     finded_code = re.findall(r"\d+", nganh_chinh_first)
        #     if len(finded_code) > 0:
        #         code = finded_code[-1]

        #     # Get the first two digits in code
        #     two_digits_code = code[:2]
        #     main_business_code = matchingKeys(code_nganh_chinh, two_digits_code)
        #     if len(main_business_code) > 0:
        #         main_business_code = main_business_code[0]

        # connection = conn()
        # cursor = connection.cursor()
        # add_sme_query = "INSERT INTO `clients` (`province`, `province_code`, `vi_name`, `en_name`, `mst`, `created_date`, `city`, `street`, `district`, `location`, `phone`, `telco`, `email`, `business`, `manager_name`, `manager_dob`, `manager_id`, `manager_addr`, `main_business`, `main_business_code`) VALUES  (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        # cursor.execute(
        #     add_sme_query,
        #     (
        #         province_zip,
        #         province_code,
        #         ten_tieng_viet,
        #         ten_tieng_anh,
        #         ma_so_thue,
        #         ngay_thanh_lap_info,
        #         dia_chi_tinh_tp,
        #         dia_chi_quan_huyen,
        #         dia_chi_phuong_xa,
        #         dia_chi_cu_the,
        #         dien_thoai,
        #         telco_name,
        #         email,
        #         nganhnghe_id,
        #         name_dai_dien,
        #         dob_dai_dien,
        #         giay_to_dai_dien,
        #         dia_chi_dai_dien,
        #         nganh_nghe_chinh,
        #         main_business_code,
        #     ),
        # )
        # connection.commit()
        # connection.close()

    except IndexError as e:
        print(e)
        next
    except Exception as e:
        print(e)
        next

def _start_ocr_file():
	# for pathdir in path_to_file:
	pathdir = path_to_file[-1]
	print(pathdir)
	datename = pathdir.split("/")[-2]
	print(datename)
	province_file_path = glob.glob(pathdir + "*/")
	index_row_csv = 0

	for provicedir in province_file_path:

		pdf_file_path = glob.glob(provicedir + "/*.pdf")
		# if provicedir.split("/")[-2] == "Đắk Nông":
		province_name = provicedir.split("/")[-2]  # Lay ten tinh/thanh pho

		province_code = [key for key in Ma_TTP if province_name in Ma_TTP[key]]

		if len(province_code) > 0:
			province_code = province_code[0]
		else:
			province_code = "..."

		province_zip = [key for key in Vung_TTP if province_name in Vung_TTP[key]]
		if len(province_zip) > 0:
			province_zip = province_zip[0]
		else:
			province_zip = "..."

		share_folder_path = ".shared/DỮ LIỆU KHDN TIỀM NĂNG 2021/"
		unit_dir = "Khu vực Miền Bắc"
		if province_zip == "MT":
			unit_dir = "Khu vực Miền Trung"
		elif province_zip == "MN":
			unit_dir = "Khu vực Miền Nam"

		province_folder = Path(PARRENT_PATH + "/temp_csv/" + share_folder_path + unit_dir + "/" + province_name)
		province_folder.mkdir(parents = True, exist_ok = True)

		# copy xlsx file
		ttdl_src_file = PARRENT_PATH + "/temp_csv/ttdl.xlsx"
		ttdl_dst_file = "{}/{}.xlsx".format(province_folder, datename)
		copyfile(ttdl_src_file, ttdl_dst_file)

		csv_file_name = datename + ".csv"
		ocr_path = province_folder.joinpath(csv_file_name)

		# xlsx_file_name = datename+ ".xlsx"
		# xlsx_path = province_folder.joinpath(xlsx_file_name)

		for path in pdf_file_path:
			index_row_csv += 1
			pre_csv_data = {
				"stt_index": index_row_csv,
				"province_code": province_code,
				"province_zip": province_zip,
			}
			if not os.path.exists(ocr_path):
				# csv_header = np.array([["KHACHHANG_ID", "MA_KH", "TEN_KH", "DIACHI_KH", "LOAIKH_ID", "KHLON_ID", "NGAYLAP_HD", "SO_GT", "NGAYCAP", "NOICAP", "HOKHAU", "SO_DT", "EMAIL", "NGUOI_DD", "NOICAP_HK", "NGAY_HK", "NGAY_SN", "DIENTHOAI_LH", "GIOITINH", "NGANHNGHE_ID", "TNC1_ID", "TNC2_ID", "TNC3_ID", "KHACHHANGCU_ID", "MA_HD", "PHANLOAIKH_ID", "LOAI", "DIEMTINNHIEM", "LOAIGT_ID", "MST"]])
				csv_header = np.array(
					[
						[
							"Số thuê bao",
							"Loại TB",
							"Tên doanh nghiệp",
							"Địa chỉ",
							"Mã Tỉnh/TP",
							"Quận/Huyện",
							"Phường/Xã",
							"Mã số thuế",
							"Ngành nghề DN",
							"Email",
							"Ngày thành lập",
							"Tên người đại diện",
							"Ngày sinh",
							"Số giấy tờ",
							"Địa chỉ thường trú",
							"Miền",
						]
					]
				)
				df = pd.DataFrame(data = csv_header)
				df.to_csv(ocr_path, index = False, header = False)

			info_arrays = parse_pdf(path)
			if len(info_arrays) > 0:
				pdf_to_csv_info = []

				# index_titles = [key for key, value in enumerate(info_arrays) if value[1] == '.' and value[2] == ' ']

				# Loai bo cac title tu Von dieu le -> Nguoi dai dien
				von_dieu_le_key_array       = [i for i, j in enumerate(info_arrays) if "6." in j and ":" in j]
				nguoi_dai_dien_key_array    = [i for i, j in enumerate(info_arrays) if "Người đại diện" in j or "8. " in j and ":" in j]
				try:
					info_company_arrays = info_arrays[: von_dieu_le_key_array[0]]
					# Xử lý các trường thông tin doanh nghiep
     
					index_titles = []
					for key, value in enumerate(info_company_arrays):
						value = value.strip()
						if len(value) > 2:
							if value[1] == "." and value[2] == " ":
								index_titles.append(key)
							elif value[-1] == ".":
								index_titles.append(key)
						elif len(value) == 2:
							if value[1] == ".":
								index_titles.append(key)

					# Đảo vị trí các index sắp xếp sai ( MST và ngày thành lập )
					MST_info = info_company_arrays[index_titles[1] - 1]
					ngay_thanh_lap_info = info_arrays[index_titles[2] - 1]

					info_company_arrays.pop(index_titles[1] - 1)
					info_company_arrays.pop(index_titles[2] - 2)

					index_titles[1] = index_titles[1] - 1
					index_titles[2] = index_titles[2] - 1

					insert_mst_at = index_titles[2] - 1
					index_ntl_at = index_titles[3] - 1
     
					info_company_arrays[insert_mst_at:insert_mst_at] = [MST_info]
					info_company_arrays[index_ntl_at:index_ntl_at] = [ngay_thanh_lap_info]

					info_company_arrays = pdf_to_array(index_titles, info_company_arrays)
					info_company_recovered = recover_info_company(info_company_arrays)

					# Xử lý các trường thông tin người đại diện

					info_director_arrays = info_arrays[nguoi_dai_dien_key_array[0] :]

					# Tạo đường dẫn tổng hợp báo cáo ngày
					report_by_day_path = Path(CSV_PATH + "/Bao cao ngay")
					report_by_day_path.mkdir(parents=True, exist_ok=True)
					report_fdate_name = datename + ".csv"
					report_fdate_path = report_by_day_path.joinpath(report_fdate_name)

					# Tạo đường dẫn tổng hợp ngành du lịch
					report_tourist_day_path = Path(CSV_PATH + "/Mkt_tourist")
					report_tourist_day_path.mkdir(parents=True, exist_ok=True)
					report_tourist_name = datename + ".csv"
					report_tourist_path = report_tourist_day_path.joinpath(
						report_tourist_name
					)

					print(ocr_path)
					parse_ocr_data(
						info_company_recovered,
						info_director_arrays,
						ocr_path,
						pre_csv_data,
						report_fdate_path,
						report_tourist_path,
					)
				except Exception as e:
					print(e)

		try:
			data = pd.read_csv(ocr_path)
			if "STT" in data:
				data.drop("STT", axis="columns").drop_duplicates(
					subset=["Mã số thuế"], keep="first", inplace=True
				)
			else:
				data.drop_duplicates(subset=["Mã số thuế"], keep="first", inplace=True)

			count_data = len(data.index)

			idx = 0
			stt_val = list(range(0, len(data)))
			data.insert(loc = idx, column = "STT", value = stt_val)
			# data.set_index('STT', inplace=True)
			# data.to_csv(ocr_path, index=False, header=True)
			# Write data to xlsx file
			start_row_raw_xlsx = 1
			append_df_to_excel(
				str(ttdl_dst_file),
				data,
				header=None,
				sheet_name="Sheet1",
				index=False,
				startrow=start_row_raw_xlsx,
			)
		except Exception as e:
			print(e)
			pass

def request_to_onesme(csv_data_report):
    sme_data = list()
    field_name = [
        "tin",
        "company",
        "address",
        "phone",
        "fax",
        "email",
        "businessId",
        "businessArea",
        "subcriberType",
        "provinceId",
        "provinceName",
        "districtId",
        "districtName",
        "wardId",
        "wardName",
        "founding",
        "representative",
        "birthday",
        "numberTax",
        "regions",
        "businessTypeName",
        "businessTypeCode",
        "businessSizeId",
        "existSme",
        "businessBlock",
        "lastName",
        "firstName",
        "taxDepartment",
        "setupAddress",
        "socialInsuranceCode",
        "nation",
        "streetName",
        "sex",
        "position",
        "nationality",
        "ethnicity",
        "identityPapers",
        "identityNumber",
        "identityDate",
        "identityAddress",
        "permanentResidence",
        "currentResidence",
        "streetId",
    ]
    for i in range(len(csv_data_report.index)):
        # for i in range(1262950, len(sme_dfs.index)):
        sme = csv_data_report.loc[i]
        # tin = sme["Ma_thue"]
        # ma_tinh = sme["Mã tỉnh"]
        # ten_tinh = sme["Tên tỉnh"]
        # ma_huyen = sme["Mã huyện"]
        # ten_huyen = sme["Tên huyện"]
        # ma_xa = sme["Mã xã"]
        # ten_xa = sme["Tên xã"]
        # company = sme["Tencs"]
        # address = sme["Dchi"]
        # phone = sme["Dthoai"]
        # fax = sme["Fax"]
        # email = sme["Email"]
        # ma_nganh = sme["Nganh_kd"]
        # ten_nganh = sme["Tennganhkd"]

        tin = sme["Mã số thuế"]
        ten_tinh = ""
        ten_huyen = sme["Quận/Huyện"]
        ten_xa = sme["Phường/Xã"]
        company = sme["Tên doanh nghiệp"]
        address = sme["Địa chỉ"]
        phone = sme["Số thuê bao"]
        email = sme["Email"]
        ten_nganh = sme["Ngành nghề DN"]
        ngay_thanh_lap = ""
        if "/" not in ngay_thanh_lap:
            ngay_thanh_lap = None
        vung = ""
        loai_thue_bao = ""
        ten_dai_dien = sme["Tên người đại diện"]
        dob = sme["Ngày sinh"]
        if "/" not in dob:
            dob = None
        giay_to = sme["Số giấy tờ"]
        diachi_thuongtru = sme["Địa chỉ thường trú"]
        diachi_lienlac = ""
        sdt_dai_dien = ""
        email_dai_dien = ""

        # tin = sme["Mã số thuế"]
        # ten_tinh = sme["Thành phố"]
        # ten_huyen = ""
        # ten_xa = ""
        # company = sme["Tên chính thức"]
        # address = sme["Địa chỉ trụ sở"]
        # phone = sme["Điện thoại/Fax"]
        # email = ""
        # ten_nganh = ""
        # ngay_thanh_lap = sme["QĐTL/Ngày cấp"]
        # if ngay_thanh_lap.count("/") != 2:
        #     ngay_thanh_lap = None
        # vung = ""
        # loai_thue_bao = ""
        # ten_dai_dien = sme["Chủ sở hữu"]
        # dob = None
        # giay_to = ""
        # diachi_thuongtru = sme["Địa chỉ chủ sở hữu"]
        # diachi_lienlac = ""
        # sdt_dai_dien = ""
        # email_dai_dien = ""

        sme_info = {
            "tin": tin,
            "company": company,
            "address": address,
            "phone": phone,
            "fax": "",
            "email": email,
            "businessId": "",
            "businessArea": ten_nganh,
            "subcriberType": "",
            "provinceId": None,
            "provinceName": ten_tinh,
            "districtId": None,
            "districtName": ten_huyen,
            "wardId": None,
            "wardName": ten_xa,
            "founding": ngay_thanh_lap,
            "representative": "",
            "birthday": dob,
            "numberTax": "",
            "regions": "",
            "businessTypeName": "",
            "businessTypeCode": None,
            "businessSizeId": None,
            "existSme": None,
            "businessBlock": "",
            "lastName": "",
            "firstName": ten_dai_dien,
            "taxDepartment": "",
            "setupAddress": "",
            "socialInsuranceCode": "",
            "nation": "",
            "streetName": diachi_lienlac,
            "sex": "",
            "position": "",
            "nationality": "",
            "ethnicity": "",
            "identityPapers": "",
            "identityNumber": giay_to,
            "identityDate": "",
            "identityAddress": diachi_thuongtru,
            "permanentResidence": "",
            "currentResidence": "",
            "streetId": None,
        }
        sme_data.append(sme_info)
        # if len(sme_data) == 10:
        headers = {"Content-type": "application/json"}
        payload = json.dumps(sme_data, ensure_ascii=False).encode("utf8")
        response = requests.post(
            "https://onesme.vn/api/enterprise/createEnterprise",
            data=payload,
            headers=headers,
        )
        if response.status_code == 200:
            try:
                sme_results = loads(response.text)
                if len(sme_results) > 0:
                    # giu lai id mst loi
                    for res in sme_results:
                        with open(current_path + "/err_file.csv", "a") as f:
                            f.write(res + "\n")
                else:
                    last_mst = sme_data[0]["tin"]
                    with open(current_path + "/success_file.csv", "a") as f:
                        f.write(last_mst + "\n")
            except Exception as e:
                print("Send ok but error get response: {}".format(e))

            sme_data.clear()
        else:
            print(response.text)
            try:
                if Path(current_path + "/err_file.csv").exists():
                    with open(current_path + "/err_file.csv", "a") as csvfile:
                        writer = csv.DictWriter(csvfile, fieldnames=field_name)
                        writer.writerows(sme_data)
                else:
                    with open(current_path + "/err_file.csv", "w") as csvfile:
                        writer = csv.DictWriter(csvfile, fieldnames=field_name)
                        writer.writeheader()
                        writer.writerows(sme_data)
            except Exception as e:
                print("Send not ok: {}".format(e))

        print("Send {} item to onesme server".format(i))

def convert_date(created_date):
    created_date = datetime.strptime(created_date, "%d/%m/%Y")
    timestamp = datetime.timestamp(created_date)
    datetime_temp = datetime.fromtimestamp(timestamp)
    created_date = datetime_temp.strftime("%Y-%m-%d")
    return created_date

def request_to_outbound(csv_data_report):
    # url = "http://10.159.22.79:8080/api/outbound/sme"
    url = "http://10.159.22.79/api/outbound/sme"

    today = datetime.today()
    ob_date = today.strftime("%Y-%m-%d")
    ob_name = "KHDN intergrate sme"

    sme_data = list()
    for i in range(len(csv_data_report.index)):
        sme = csv_data_report.loc[i]
        so_thue_bao = sme["Số thuê bao"]
        loai_thue_bao = sme["Loại TB"]
        ten_dn = sme["Tên doanh nghiệp"]
        dia_chi = sme["Địa chỉ"]
        ma_tinh = sme["Mã Tỉnh/TP"]
        quan_huyen = sme["Quận/Huyện"]
        phuong_xa = sme["Phường/Xã"]
        mst = sme["Mã số thuế"]
        nganh_nghe = sme["Ngành nghề DN"]
        email = sme["Email"]
        created_date = sme["Ngày thành lập"]
        try:
            if "/" in created_date:
                created_date = convert_date(created_date)
            else:
                created_date = ""
        except Exception as e:
            created_date = ""
            print(e)

        name_dd = sme["Tên người đại diện"]
        dob = sme["Ngày sinh"]
        try:
            if "/" in dob:
                dob = convert_date(dob)
            else:
                dob = ""
        except Exception as e:
            dob = ""
            print(e)

        so_giay_to = sme["Số giấy tờ"]
        dia_chi_dd = sme["Địa chỉ thường trú"]
        mien = sme["Miền"]

        sme_info = {
            "so_thue_bao": so_thue_bao,
            "loai_thue_bao": loai_thue_bao,
            "ten_doanh_nghiep": ten_dn,
            "dia_chi": dia_chi,
            "ma_tinh": ma_tinh,
            "quan_huyen": quan_huyen,
            "phuong_xa": phuong_xa,
            "ma_so_thue": mst,
            "nganh_nghe": nganh_nghe,
            "email": email,
            "ngay_thanh_lap": created_date,
            "ten_nguoi_dai_dien": name_dd,
            "ngay_sinh": dob,
            "so_giay_to": so_giay_to,
            "dia_chi_thuong_chu": dia_chi_dd,
            "mien": mien,
        }
        sme_data.append(sme_info)

    payload = json.dumps(
        {
            "username": "sme_api",
            "password": "sme_api@12356",
            "data_date": ob_date,
            "ob_name": ob_name,
            "data": sme_data,
        },
        ensure_ascii=False,
    ).encode("utf8")
    headers = {"Content-Type": "application/json"}

    response = requests.request("POST", url, headers=headers, data=payload)
    resp = response.json()
    try:
        resp_code = resp["Code"]
        if resp_code != 1:
            resp_data = resp["Data"]
            for data in resp_data:
                with open(current_path + "/err_file_outbound.txt", "a") as f:
                    f.write(
                        "Date: {} - SDT: {} - error {}".format(
                            ob_date, data["so_thue_bao"], data["error_message"]
                        )
                        + "\n"
                    )
                f.close()
    except Exception as e:
        print("Error parse response json: {}".format(e))

    print(response.text)

# Request to khdn

def no_accent_vietnamese(s):
    s = re.sub('[áàảãạăắằẳẵặâấầẩẫậ]', 'a', s)
    s = re.sub('[ÁÀẢÃẠĂẮẰẲẴẶÂẤẦẨẪẬ]', 'A', s)
    s = re.sub('[éèẻẽẹêếềểễệ]', 'e', s)
    s = re.sub('[ÉÈẺẼẸÊẾỀỂỄỆ]', 'E', s)
    s = re.sub('[óòỏõọôốồổỗộơớờởỡợ]', 'o', s)
    s = re.sub('[ÓÒỎÕỌÔỐỒỔỖỘƠỚỜỞỠỢ]', 'O', s)
    s = re.sub('[íìỉĩị]', 'i', s)
    s = re.sub('[ÍÌỈĨỊ]', 'I', s)
    s = re.sub('[úùủũụưứừửữự]', 'u', s)
    s = re.sub('[ÚÙỦŨỤƯỨỪỬỮỰ]', 'U', s)
    s = re.sub('[ýỳỷỹỵ]', 'y', s)
    s = re.sub('[ÝỲỶỸỴ]', 'Y', s)
    s = re.sub('đ', 'd', s)
    s = re.sub('Đ', 'D', s)
    return s

def get_location_name(s):
    s = re.sub(r'^thanh pho|^quan|^huyen|^thi xa|^xa|^phuong', '', s)
    return s

def get_tinh_id(ma_tinh):
    tinh_id = Khdn_tinh_id[ma_tinh]
    # if ma_tinh == "HUE" or ma_tinh == "KTM":
    #     ten_tinhs = ten_tinh.split("-")
    #     ten_tinh_short = ten_tinh[-1:]
    # else:
    #     ten_tinhs = ten_tinh.split(" ")
    #     ten_tinh_short = (" ").join(ten_tinhs[-2:])
    # tinh_id = Khdn_tinh_id[ten_tinh_short.lower()]
    return tinh_id 

def get_quan_id(tinh_id, ten_quan):
    quan_id = 0
    quan_df = None
    quan_file_path = khdn_quan_path + "/{}.csv".format(tinh_id)
    try:
        quan_df = pd.read_csv(quan_file_path)
    except Exception as e:
        print("File khdn_quan_path not found: {}".format(khdn_quan_path))
    if quan_df is not None:
        ten_quan = no_accent_vietnamese(ten_quan)
        ten_quan = ten_quan.lower()
        ten_quan = get_location_name(ten_quan)
        ten_quan = ten_quan.strip()
        quan_info = quan_df[quan_df['NAME'].str.contains(ten_quan)]

        if len(quan_info) > 0:
            quan_id = quan_info.iloc[0]['ID']
    return quan_id, ten_quan

def get_phuong_id(ten_quan, ten_phuong):
    file_phuong_path = ""
    phuong_id = 0
    phuong_df = None
    for quan_path in glob.glob(khdn_phuong_path + "/*"):
        ten_quan_path = quan_path.split("/")[-1]
        if ten_quan in ten_quan_path:
            file_phuong_path = os.path.join(khdn_phuong_path,ten_quan_path)
    
    if file_phuong_path != "":
        try:
            phuong_df = pd.read_csv(file_phuong_path)
        except Exception as e:
            print("File khdn_quan_path not found: {}".format(khdn_phuong_path))
        if phuong_df is not None:
            ten_phuong = no_accent_vietnamese(ten_phuong)
            ten_phuong = ten_phuong.lower()
            ten_phuong = get_location_name(ten_phuong)
            ten_phuong = ten_phuong.strip()
            phuong_info = phuong_df[phuong_df['NAME'].str.contains(ten_phuong)]
            if len(phuong_info) > 0:
                phuong_id = phuong_info.iloc[0]['ID']
    return phuong_id

def request_to_khdn(csv_data_report):
    url = "http://10.156.4.137:5000/mapi/services/new_khdn"

    for i in range(len(csv_data_report.index)):
        sme = csv_data_report.loc[i]

        so_thue_bao = sme["Số thuê bao"]
        ten_dn = sme["Tên doanh nghiệp"]
        dia_chi = sme["Địa chỉ"]
        ma_tinh = get_tinh_id(sme["Mã Tỉnh/TP"])
        quan_huyen, ten_quan = get_quan_id(ma_tinh, sme["Quận/Huyện"])
        phuong_xa = get_phuong_id(ten_quan, sme["Phường/Xã"])

        mst = sme["Mã số thuế"]
        email = sme["Email"]
        created_date = sme["Ngày thành lập"]
        try:
            if "/" not in created_date:
                created_date = ""
        except Exception as e:
            created_date = ""
            print(e)

        name_dd = sme["Tên người đại diện"]
        dob = sme["Ngày sinh"]
        try:
            if "/" not in dob:
                dob = ""
        except Exception as e:
            dob = ""
            print(e)

        so_giay_to = sme["Số giấy tờ"]
        country = "1"

        sme_info = {
            "full_name": ten_dn,
            "short_name": "",
            "full_name_en": "",
            "mst": mst[1:],
            "ma_ngan_sach": "",
            "addr": dia_chi,
            "country": country,
            "province": ma_tinh,
            "district": "{}".format(quan_huyen),
            "town": "{}".format(phuong_xa),
            "village": "",
            "loai_gt": "1",
            "dd_phone": "",
            "dd_email": email,
            "ngay_thanh_lap": created_date,
            "dd_name": name_dd,
            "phone": so_thue_bao,
            "dd_birthday": dob,
            "so_gt": so_giay_to,
            "loai_hinh_to_chuc": "1",
            "dd_gender": "",
            "dd_sothich": "",
            "line_kh": "5",
            "phan_loai": "6",
            "phan_hang": "",
            "sub_line": "15",
            "ke_toan_truong": "",
            "phone_ktt": "",
            "email_ktt": "",
            "nganh_nghe": "",
            "ma_nganh": "",
            "so_luong_laodong": "",
            "so_luong_chinhanh": "",
            "quy_mo_kh": "",
            "so_tk": "",
            "ngan_hang": "",
            "ma_khtn": "",
            "loai_dn": "",
            "doanh_thu": "",
            "ten_congty_me": "",
        }
        payload = json.dumps(sme_info, ensure_ascii=False).encode("utf8")
        headers = {"Content-Type": "application/json"}

        response = requests.request("POST", url, headers=headers, data=payload)
        resp = response.json()
        try:
            resp_code = resp["Code"]
            if resp_code != 1:
                with open(current_path + "/err_file_khdn.txt", "a") as f:
                        f.write("- MST: {} - resp_code {}".format(
                                mst, resp_code
                            )
                            + "\n"
                        )
                f.close()
        except Exception as e:
            print("Error parse response json: {}".format(e))
            continue

        print(resp)
    
def convert_created_date(ngay_thanh_lap_info):
    if "/" in ngay_thanh_lap_info:
        ngay_thanh_lap_info = datetime.strptime(ngay_thanh_lap_info, "%d/%m/%Y")
        timestamp = datetime.timestamp(ngay_thanh_lap_info)
        datetime_temp = datetime.fromtimestamp(timestamp)
        ngay_thanh_lap_info = datetime_temp.strftime("%Y-%m-%d %H:%M:%S")
    else:
        ngay_thanh_lap_info = ""
    return ngay_thanh_lap_info

def get_main_business(nganh_nghe):
    nganh_nghe_list = nganh_nghe.split("\n")
    search_nn = [x for x in nganh_nghe_list if "(chính)" in x.lower()]
    nganh_nghe_chinh = "\n".join(search_nn)
    return nganh_nghe_chinh

def get_key_from_business(businesses):
    split_text = businesses.split(" ")
    matched_result = ""
    for text in split_text:
        if (("(chính)") in text.lower()) or (text.isdigit() and len(text) == 4):
            match_key = matchingKeys(code_nganh_chinh, text[:2])
            if len(match_key) > 0:
                return match_key[0]
            else:
                return matched_result

def get_type_sme(name):
    result = "khác"
    if name != "":
        if len([x for x in ["tnhh", "trách nhiệm hữu hạn"] if x in name]) > 0:
            result = "TNHH MTV"
        elif (
            len(
                [x for x in [
                        "ctcp",
                        "cổ phần",
                        "cô phân",
                        "cp",
                        "CÔ PHẦN",
                        "CÔ PHÁN",
                        "CÔ PHẢN",
                        "CÔ PHÁẢN",
                        "CÔ PHAN",
                        "CÔ PHÁÂN",
                    ] if x.lower() in name]) > 0):
            result = "CTCP"
        elif (
            len([x for x in [
                        "Doanh Nghiệp Tư Nhân",
                        "DNTN",
                        "Doanh Nghiệp TN",
                    ] if x.lower() in name ]) > 0):
            result = "DNTN"
        elif (
            len([x for x in [
                        "hợp danh",
                    ] if x.lower() in name]) > 0):
            result = "HD"
    return result

def update_to_mysql(csv_data_report):
    # Convert created date
    csv_data_report["created_date"] = csv_data_report["Ngày thành lập DN"].apply(
        lambda x: convert_created_date(x)
    )
    csv_data_report["main_business"] = csv_data_report[
        "Lĩnh vực (Ngành nghề kinh doanh chính)"
    ].apply(lambda x: get_main_business(x))

    csv_data_report["main_business_code"] = csv_data_report["main_business"].apply(
        lambda x: get_key_from_business(x)
    )

    csv_data_report["enterprise_type"] = csv_data_report["Tên Tiếng Việt"].apply(
        lambda x: get_type_sme(x.lower())
    )

    csv_data_report.rename(
        columns={
            "Vùng": "province",
            "Mã T/TP": "province_code",
            "Tên Tiếng Việt": "vi_name",
            "Tên Tiếng Anh": "en_name",
            "Mã số thuế/Mã ngân sách": "mst",
            "Tỉnh/TP": "city",
            "Quận/Huyện": "street",
            "Phường/Xã": "district",
            "Địa chỉ cụ thể ": "location",
            "Điện thoại": "phone",
            "CHECK VINA": "telco",
            "Lĩnh vực (Ngành nghề kinh doanh chính)": "business",
            "Họ và tên.1": "manager_name",
            "Ngày sinh.1": "manager_dob",
            "Số giấy tờ.1": "manager_id",
            "Địa chỉ thường trú.1": "manager_addr",
            "Email": "email",
        },
        inplace=True,
    )

    csv_data_report.drop(
        [
            "Ngày thành lập DN",
            "Email.1",
            "Họ và tên",
            "Ngày sinh",
            "Số giấy tờ",
            "Địa chỉ thường trú",
            "Địa chỉ liên lạc.1",
            "Số điện thoại.1",
            "Địa chỉ liên lạc",
            "Email.2",
            "Số điện thoại",
        ],
        axis=1,
        inplace=True,
    )
    query_origin = "SELECT DISTINCT mst FROM clients"
    mst_existed = pd.read_sql(query_origin, con=sqlEngine)
    df_all = csv_data_report.merge(mst_existed, on=["mst"], how="left", indicator=True)
    df_all = df_all[df_all["_merge"] == "left_only"]
    df_all.drop(['_merge'], axis=1, inplace=True)

    with sqlEngine.begin() as connection:
        df_all.to_sql(
            name="clients", con=connection, if_exists="append", index=False
        )
    
    # Xoa file temp
    # if os.path.isfile(temp_all_ptdl_path):
    #     os.remove(temp_all_ptdl_path)
    # else:  ## Show an error ##
    #     print("Error: %s file not found" % temp_all_ptdl_path)

def remove_split(number):
    result = number
    if result != None:
        if "\n" in number:
            result = number.split("\n")[0]
        if "," in number:
            result = number.split(",")[0]
        if ";" in number:
            result = number.split(";")[0]
        if "-" in number:
            result = number.split("-")[0]
        if "." in number:
            result = number.split(".")[0]
    return result

def remove_bracket(number):
    result = number
    if result != None:
        if "(" in number:
            result = re.sub(r"\([^)]*\)", "", number)
    return result

def convert_phone_number(phone):
    # remove special characters
    phone = phone.apply(lambda x: remove_split(x) if x != None else "")
    # remove plus
    phone = phone.apply(lambda x: remove_bracket(x) if x != None else "")
    # remove long len
    phone = phone.apply(lambda x: x if not "@" in x else "")
    return phone

def _update_data_other_sys(type_sys):
	if type_sys == TYPE_KHDN:
		query_data = "SELECT phone AS `Số thuê bao`, telco AS `Loại TB`, vi_name AS `Tên doanh nghiệp`, province_code AS `Mã Tỉnh/TP`, location AS `Địa chỉ`, street AS `Quận/Huyện`, district AS `Phường/Xã`, mst AS `Mã số thuế`, email AS `Email`, created_date AS `Ngày thành lập`, manager_name AS `Tên người đại diện`, manager_dob AS `Ngày sinh`, manager_id AS `Số giấy tờ` FROM clients WHERE created_date > '2022-09-01' AND province_code <> '...'"
		csv_data_report = pd.read_sql(query_data, con=sqlEngine)	
		csv_data_report = csv_data_report.applymap(str)
		csv_data_report["Số thuê bao"] = convert_phone_number(
			csv_data_report["Số thuê bao"]
		)
		csv_data_report.fillna("", inplace=True)	
		csv_data_report["Mã số thuế"] = csv_data_report["Mã số thuế"].str.lstrip("'")
		print(csv_data_report.head())
		# request_to_khdn(csv_data_report)

if __name__ == '__main__':
    # path = "/Users/dinhvan/Documents/Projects/OCR/Crawl_info/temp/2022_06_01/An Giang/01_06_2022_1_0.pdf"
    # parse_pdf(path)
	_start_ocr_file()
	# Tổng hợp báo cáo toàn bộ
	report_by_day_path = Path(CSV_PATH + "/Bao cao ngay")
	if os.path.exists(report_by_day_path):
		all_report_files = sorted(glob.glob(str(report_by_day_path) + "/*.csv"))
		if len(all_report_files) > 0:
			print("last file date: {}".format(all_report_files[-1]))
			# all_report_path = Path(CSV_PATH).joinpath("all_reports_new_1.csv")
			csv_data_report = pd.read_csv(all_report_files[-1], sep="\t", encoding="utf-8")
			csv_data_report = csv_data_report.applymap(str)
			csv_data_report["Số thuê bao"] = convert_phone_number(
				csv_data_report["Số thuê bao"]
			)
			csv_data_report.fillna("", inplace=True)
			request_to_onesme(csv_data_report)
			request_to_outbound(csv_data_report)
			request_to_khdn(csv_data_report)

	# # Update all_ptdl to mysql
	# if os.path.exists(os.path.join(PARRENT_PATH, "temp_ptdl.csv")):
	# 	temp_all_ptdl_path = os.path.join(PARRENT_PATH, "temp_ptdl.csv")
	# 	csv_ptdl_data = pd.read_csv(temp_all_ptdl_path, sep=",", encoding="utf-8")
	# 	csv_ptdl_data = csv_ptdl_data[csv_ptdl_data['Mã số thuế/Mã ngân sách'].str.contains('lập:') == False]
	# 	csv_ptdl_data.fillna("", inplace=True)
	# 	update_to_mysql(csv_ptdl_data)
	
	# _update_data_other_sys(type_sys=TYPE_KHDN)



